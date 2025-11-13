"""
Automobile Manufacturing Financial Platform - Streamlit Web Interface
Interactive labor management, CAPEX scheduling, financial modeling
Author: Advanced Analytics Team
Version: 1.0 (November 2025)
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from copy import deepcopy
from typing import Dict, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import platform modules
from financial_model import run_financial_model, CompanyConfig
from labor_management import (
    initialize_default_labor_structure, LaborScheduleManager, LaborCostSchedule,
    ProductionLinkedLabor, LaborType, EmploymentStatus, JobCategory
)
from capex_management import initialize_default_capex, CapexScheduleManager

# =====================================================
# EXCEL EXPORT HELPER FUNCTIONS
# =====================================================

def _generate_excel_bytes(model: Dict, labor_manager, capex_manager, scenario: str = "Base Case") -> bytes:
    """
    Generate Excel workbook with financial model, labor schedule, and CAPEX schedule
    
    Args:
        model: Financial model results dictionary
        labor_manager: Labor management object
        capex_manager: CAPEX management object
        scenario: Scenario name for the model
    
    Returns:
        Bytes of Excel workbook
    """
    if not OPENPYXL_AVAILABLE:
        st.error("openpyxl is required for Excel export. Install with: pip install openpyxl")
        return None
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_format = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'
    number_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== SHEET 1: EXECUTIVE SUMMARY ==========
    ws_summary = wb.create_sheet("Executive Summary", 0)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20
    
    row = 1
    ws_summary[f'A{row}'] = f"Automobile Manufacturing Financial Model - {scenario}"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    ws_summary[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    row += 2
    
    # Summary metrics
    ws_summary[f'A{row}'] = "Key Metrics"
    ws_summary[f'A{row}'].font = subtitle_font
    row += 1
    
    metrics = [
        ("2026 Revenue", model['revenue'][0] if isinstance(model['revenue'], list) else model['revenue']),
        ("2026 EBIT", model['ebit'][0] if isinstance(model['ebit'], list) else model['ebit']),
        ("2026 Net Profit", model['net_profit'][0] if isinstance(model['net_profit'], list) else model['net_profit']),
        ("2026 FCF", model['fcf'][0] if isinstance(model['fcf'], list) else model['fcf']),
        ("Enterprise Value", model.get('enterprise_value', 0)),
        ("5-Year Total FCF", sum(model['fcf']) if isinstance(model['fcf'], list) else model['fcf']),
    ]
    
    for label, value in metrics:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'B{row}'].number_format = currency_format
        ws_summary[f'B{row}'].border = thin_border
        row += 1
    
    # ========== SHEET 2: FINANCIAL MODEL ==========
    ws_financial = wb.create_sheet("Financial Model", 1)
    
    # Header
    headers = ['Year', 'Revenue', 'COGS', 'Gross Profit', 'OPEX', 'EBITDA', 'Depreciation', 'EBIT', 'Interest', 'EBT', 'Taxes', 'Net Profit', 'FCF']
    for col, header in enumerate(headers, 1):
        cell = ws_financial.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Data rows
    years = list(range(2026, 2031))
    for row_idx, year in enumerate(years, 2):
        ws_financial.cell(row=row_idx, column=1, value=year)
        ws_financial.cell(row=row_idx, column=2, value=model['revenue'][row_idx-2] if isinstance(model['revenue'], list) else model['revenue'])
        ws_financial.cell(row=row_idx, column=3, value=model['cogs'][row_idx-2] if isinstance(model['cogs'], list) else model['cogs'])
        ws_financial.cell(row=row_idx, column=4, value=model['revenue'][row_idx-2] - model['cogs'][row_idx-2] if isinstance(model['revenue'], list) else model['revenue'] - model['cogs'])
        ws_financial.cell(row=row_idx, column=5, value=model['opex'][row_idx-2] if isinstance(model['opex'], list) else model['opex'])
        ws_financial.cell(row=row_idx, column=6, value=model['revenue'][row_idx-2] - model['cogs'][row_idx-2] - model['opex'][row_idx-2] if isinstance(model['revenue'], list) else model['revenue'] - model['cogs'] - model['opex'])
        ws_financial.cell(row=row_idx, column=7, value=model.get('depreciation', [0]*5)[row_idx-2])
        ws_financial.cell(row=row_idx, column=8, value=model['ebit'][row_idx-2] if isinstance(model['ebit'], list) else model['ebit'])
        ws_financial.cell(row=row_idx, column=9, value=0)  # Interest
        ws_financial.cell(row=row_idx, column=10, value=model['ebit'][row_idx-2] if isinstance(model['ebit'], list) else model['ebit'])  # EBT
        ws_financial.cell(row=row_idx, column=11, value=(model['ebit'][row_idx-2] * 0.21) if isinstance(model['ebit'], list) else (model['ebit'] * 0.21))  # Taxes
        ws_financial.cell(row=row_idx, column=12, value=model['net_profit'][row_idx-2] if isinstance(model['net_profit'], list) else model['net_profit'])
        ws_financial.cell(row=row_idx, column=13, value=model['fcf'][row_idx-2] if isinstance(model['fcf'], list) else model['fcf'])
        
        # Format as currency
        for col in range(2, 14):
            ws_financial.cell(row=row_idx, column=col).number_format = currency_format
            ws_financial.cell(row=row_idx, column=col).border = thin_border
    
    # Set column widths
    ws_financial.column_dimensions['A'].width = 12
    for col in range(2, 14):
        ws_financial.column_dimensions[get_column_letter(col)].width = 16
    
    # ========== SHEET 3: LABOR SCHEDULE ==========
    ws_labor = wb.create_sheet("Labor Schedule", 2)
    
    try:
        cost_schedule = LaborCostSchedule(labor_manager)
        labor_df = cost_schedule.generate_5year_schedule(salary_growth=st.session_state.salary_growth_rate)
        
        # Write headers
        for col, header in enumerate(labor_df.columns, 1):
            cell = ws_labor.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, row_data in enumerate(labor_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_labor.cell(row=row_idx, column=col_idx, value=value)
                if col_idx > 1:  # Format numeric columns as currency or number
                    if 'Cost' in labor_df.columns[col_idx-1]:
                        cell.number_format = currency_format
                    else:
                        cell.number_format = number_format
                cell.border = thin_border
        
        # Auto-size columns
        for col_idx, col_header in enumerate(labor_df.columns, 1):
            ws_labor.column_dimensions[get_column_letter(col_idx)].width = 18
    
    except Exception as e:
        ws_labor['A1'] = f"Error generating labor schedule: {str(e)}"
    
    # ========== SHEET 4: CAPEX SCHEDULE ==========
    ws_capex = wb.create_sheet("CAPEX Schedule", 3)
    
    try:
        capex_items = capex_manager.list_items()
        
        headers = ['Item ID', 'Asset Name', 'Category', 'Acquisition Cost', 'Acquisition Date', 'Useful Life (Years)', 'Annual Depreciation']
        for col, header in enumerate(headers, 1):
            cell = ws_capex.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Write CAPEX items
        for row_idx, item in enumerate(capex_items, 2):
            ws_capex.cell(row=row_idx, column=1, value=item.item_id).border = thin_border
            ws_capex.cell(row=row_idx, column=2, value=item.name).border = thin_border
            ws_capex.cell(row=row_idx, column=3, value=item.category).border = thin_border
            
            cost_cell = ws_capex.cell(row=row_idx, column=4, value=item.amount)
            cost_cell.number_format = currency_format
            cost_cell.border = thin_border
            
            ws_capex.cell(row=row_idx, column=5, value=item.acquisition_date).border = thin_border
            ws_capex.cell(row=row_idx, column=6, value=item.useful_life).border = thin_border
            
            annual_dep = item.amount / item.useful_life if item.useful_life > 0 else 0
            dep_cell = ws_capex.cell(row=row_idx, column=7, value=annual_dep)
            dep_cell.number_format = currency_format
            dep_cell.border = thin_border
        
        # Set column widths
        ws_capex.column_dimensions['A'].width = 12
        ws_capex.column_dimensions['B'].width = 20
        ws_capex.column_dimensions['C'].width = 15
        ws_capex.column_dimensions['D'].width = 18
        ws_capex.column_dimensions['E'].width = 18
        ws_capex.column_dimensions['F'].width = 18
        ws_capex.column_dimensions['G'].width = 20
    
    except Exception as e:
        ws_capex['A1'] = f"Error generating CAPEX schedule: {str(e)}"
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# =====================================================
# PAGE CONFIG
# ====================================================="

st.set_page_config(
    page_title="Manufacturing Financial Platform",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =====================================================
# SESSION STATE INITIALIZATION
# =====================================================

def initialize_session_state():
    """Initialize or restore session state"""
    if 'labor_manager' not in st.session_state:
        st.session_state.labor_manager = initialize_default_labor_structure()
    
    if 'capex_manager' not in st.session_state:
        st.session_state.capex_manager = initialize_default_capex(CapexScheduleManager())
    
    if 'financial_model' not in st.session_state:
        st.session_state.financial_model = None
    
    if 'salary_growth_rate' not in st.session_state:
        st.session_state.salary_growth_rate = 0.05
    
    if 'last_update' not in st.session_state:
        st.session_state.last_update = None
    
    if 'excel_bytes_map' not in st.session_state:
        st.session_state.excel_bytes_map = {}
    
    if 'selected_scenario' not in st.session_state:
        st.session_state.selected_scenario = "Base Case"

initialize_session_state()

# =====================================================
# SIDEBAR - NAVIGATION & SETTINGS
# =====================================================

with st.sidebar:
    st.markdown("# ⚙️ Platform Settings")
    
    # Navigation
    page = st.radio(
        "Select Module:",
        ["🏠 Dashboard", "👥 Labor Management", "🏗️ CAPEX Management", 
         "💰 Financial Model", "📈 Reports"],
        help="Choose which module to work with"
    )
    
    st.divider()
    
    # Global Settings
    st.markdown("### 📋 Global Parameters")
    
    salary_growth = st.slider(
        "Annual Salary Growth Rate (%)",
        min_value=0,
        max_value=10,
        value=int(st.session_state.salary_growth_rate * 100),
        help="Applied to all labor cost projections"
    )
    st.session_state.salary_growth_rate = salary_growth / 100
    
    st.divider()
    
    # Platform Info
    st.markdown("### ℹ️ Platform Info")
    st.info(
        "**Manufacturing Financial Platform v1.0**\n\n"
        "• Labor management (CRUD)\n"
        "• CAPEX scheduling\n"
        "• Financial modeling\n"
        "• Advanced reporting\n\n"
        "*Select module to navigate*"
    )

# =====================================================
# PAGE 1: DASHBOARD
# =====================================================

if page == "🏠 Dashboard":
    st.markdown("# 📊 Executive Dashboard")
    
    # Run financial model
    cfg = CompanyConfig(
        labor_manager=st.session_state.labor_manager,
        capex_manager=st.session_state.capex_manager
    )
    model = run_financial_model(cfg)
    st.session_state.financial_model = model
    
    # Key Metrics
    st.markdown("## 📈 Key Performance Indicators")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        revenue_2026 = model['revenue'][0]
        st.metric("2026 Revenue", f"${revenue_2026/1e6:.1f}M", delta="Year 1")
    
    with col2:
        profit_2026 = model['net_profit'][0]
        margin = (profit_2026/revenue_2026*100) if revenue_2026 > 0 else 0
        st.metric("2026 Net Profit", f"${profit_2026/1e6:.1f}M", delta=f"{margin:.1f}%")
    
    with col3:
        headcount = st.session_state.labor_manager.get_total_headcount(2026)
        st.metric("Total Headcount", f"{headcount} employees", delta="Current")
    
    with col4:
        ev = model['enterprise_value']
        st.metric("Enterprise Value", f"${ev/1e6:.1f}M", delta="DCF")
    
    # Financial Overview
    st.markdown("## 💼 5-Year Financial Forecast")
    
    col1, col2 = st.columns(2)
    
    with col1:
        forecast_df = pd.DataFrame({
            'Year': range(2026, 2031),
            'Revenue ($M)': [x/1e6 for x in model['revenue']],
            'EBIT ($M)': [x/1e6 for x in model['ebit']],
            'Net Profit ($M)': [x/1e6 for x in model['net_profit']],
            'FCF ($M)': [x/1e6 for x in model['fcf']]
        })
        st.dataframe(forecast_df, use_container_width=True, hide_index=True)
    
    with col2:
        fig = px.line(
            forecast_df,
            x='Year',
            y=['Revenue ($M)', 'Net Profit ($M)'],
            markers=True,
            title="Revenue & Profit Forecast",
            labels={'value': 'Amount ($M)'}
        )
        st.plotly_chart(fig, use_container_width=True)
    
    # Labor & CAPEX Overview
    st.markdown("## 👥 Workforce & Assets")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        labor_costs = st.session_state.labor_manager.get_labor_cost_by_type(2026, st.session_state.salary_growth_rate)
        st.markdown("### Labor Costs (2026)")
        st.metric("Direct Labor", f"${labor_costs['Direct']/1e6:.2f}M")
        st.metric("Indirect Labor", f"${labor_costs['Indirect']/1e6:.2f}M")
        st.metric("Total Labor", f"${(labor_costs['Direct'] + labor_costs['Indirect'])/1e6:.2f}M")
    
    with col2:
        hc_types = st.session_state.labor_manager.get_headcount_by_type(2026)
        st.markdown("### Headcount (2026)")
        st.metric("Direct Labor HC", f"{hc_types['Direct']} employees")
        st.metric("Indirect Labor HC", f"{hc_types['Indirect']} employees")
        st.metric("Total HC", f"{hc_types['Direct'] + hc_types['Indirect']} employees")
    
    with col3:
        capex_items = st.session_state.capex_manager.list_items()
        total_capex = st.session_state.capex_manager.total_capex()
        st.markdown("### Capital Assets (2026)")
        st.metric("Total CAPEX", f"${total_capex/1e6:.2f}M")
        st.metric("# Assets", f"{len(capex_items)}")
        st.metric("Annual Depreciation", f"${model['depreciation'][0]/1e3:.0f}K")

# =====================================================
# PAGE 2: LABOR MANAGEMENT
# =====================================================

elif page == "👥 Labor Management":
    st.markdown("# 👥 Labor Position Management")
    
    # Create tabs
    tab1, tab2, tab3 = st.tabs(["📋 Current Positions", "➕ Add Position", "✏️ Edit Position"])
    
    # TAB 1: View Positions
    with tab1:
        st.markdown("## Current Labor Schedule")
        
        positions_df = st.session_state.labor_manager.get_position_summary()
        st.dataframe(positions_df, use_container_width=True, hide_index=True)
        
        # Summary statistics
        col1, col2, col3 = st.columns(3)
        
        labor_costs = st.session_state.labor_manager.get_labor_cost_by_type(2026, st.session_state.salary_growth_rate)
        total_cost = labor_costs['Direct'] + labor_costs['Indirect']
        hc = st.session_state.labor_manager.get_total_headcount(2026)
        
        with col1:
            st.metric("Total Headcount", hc)
        with col2:
            st.metric("Total Annual Cost", f"${total_cost/1e6:.2f}M")
        with col3:
            st.metric("Cost per Employee", f"${total_cost/hc:,.0f}" if hc > 0 else "N/A")
        
        # Labor Cost Schedule
        st.markdown("## 5-Year Labor Cost Schedule")
        
        cost_schedule = LaborCostSchedule(st.session_state.labor_manager)
        schedule_df = cost_schedule.generate_5year_schedule(salary_growth=st.session_state.salary_growth_rate)
        
        # Format for display
        display_df = schedule_df.copy()
        for col in ['Direct Labor Cost', 'Indirect Labor Cost', 'Total Labor Cost']:
            display_df[col] = display_df[col].apply(lambda x: f"${x/1e6:.2f}M")
        
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        # Visualization
        fig = px.line(
            schedule_df,
            x='Year',
            y=['Direct Labor Cost', 'Indirect Labor Cost'],
            markers=True,
            title="Labor Cost Forecast (2026-2030)",
            labels={'value': 'Cost ($)', 'variable': 'Labor Type'}
        )
        st.plotly_chart(fig, use_container_width=True)
    
    # TAB 2: Add Position
    with tab2:
        st.markdown("## Add New Labor Position")
        
        col1, col2 = st.columns(2)
        
        with col1:
            position_name = st.text_input("Position Name", value="New Position")
            labor_type = st.selectbox("Labor Type", [LaborType.DIRECT, LaborType.INDIRECT])
            job_category = st.selectbox("Job Category", list(JobCategory))
            headcount = st.number_input("Headcount", min_value=1, value=1, step=1)
            annual_salary = st.number_input("Annual Salary ($)", min_value=20000, value=40000, step=5000)
        
        with col2:
            status = st.selectbox("Employment Status", list(EmploymentStatus))
            start_year = st.number_input("Start Year", min_value=2026, value=2026, step=1)
            benefits_percent = st.slider("Benefits % of Salary", 0.0, 1.0, 0.25, step=0.05)
            overtime_hours = st.number_input("Annual Overtime Hours", min_value=0, value=0, step=50)
            training_cost = st.number_input("Annual Training Cost ($)", min_value=0, value=0, step=500)
        
        if st.button("✅ Add Position"):
            try:
                position_id = st.session_state.labor_manager.add_position(
                    position_name=position_name,
                    labor_type=labor_type,
                    job_category=job_category,
                    headcount=headcount,
                    annual_salary=annual_salary,
                    status=status,
                    start_year=start_year,
                    benefits_percent=benefits_percent,
                    overtime_hours_annual=overtime_hours,
                    training_cost=training_cost
                )
                st.success(f"✅ Position added! ID: {position_id}")
                st.session_state.last_update = datetime.now()
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    # TAB 3: Edit Position
    with tab3:
        st.markdown("## Edit Labor Position")
        
        positions = st.session_state.labor_manager.get_all_positions()
        if not positions:
            st.warning("No positions to edit")
        else:
            pos_display = {f"{p.position_id} - {p.position_name}": p.position_id for p in positions}
            selected_display = st.selectbox("Select Position", list(pos_display.keys()))
            selected_id = pos_display[selected_display]
            pos = st.session_state.labor_manager.get_position(selected_id)
            
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Position Name", value=pos.position_name)
                new_headcount = st.number_input("Headcount", min_value=0, value=pos.headcount)
                new_salary = st.number_input("Annual Salary", min_value=0, value=int(pos.annual_salary))
            
            with col2:
                new_status = st.selectbox("Status", list(EmploymentStatus), index=list(EmploymentStatus).index(pos.status))
                new_benefits = st.slider("Benefits %", 0.0, 1.0, pos.benefits_percent)
                new_training = st.number_input("Training Cost", min_value=0, value=int(pos.training_cost_annual))
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("✅ Save Changes"):
                    try:
                        st.session_state.labor_manager.edit_position(
                            selected_id,
                            position_name=new_name,
                            headcount=new_headcount,
                            annual_salary=new_salary,
                            status=new_status,
                            benefits_percent=new_benefits,
                            training_cost_annual=new_training
                        )
                        st.success("✅ Position updated!")
                        st.session_state.last_update = datetime.now()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
            
            with col2:
                if st.button("🗑️ Remove Position"):
                    try:
                        st.session_state.labor_manager.remove_position(selected_id)
                        st.success("✅ Position removed!")
                        st.session_state.last_update = datetime.now()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")

# =====================================================
# PAGE 3: CAPEX MANAGEMENT
# =====================================================

elif page == "🏗️ CAPEX Management":
    st.markdown("# 🏗️ Capital Expenditure Management")
    
    tab1, tab2, tab3 = st.tabs(["📋 Current Assets", "➕ Add Asset", "✏️ Edit Asset"])
    
    # TAB 1: View Assets
    with tab1:
        st.markdown("## Current CAPEX Schedule")
        
        assets = st.session_state.capex_manager.list_items()
        items_data = []
        for item in assets:
            items_data.append({
                'ID': item.item_id,
                'Name': item.name,
                'Category': item.category,
                'Amount ($M)': f"${item.amount/1e6:.2f}",
                'Life (years)': item.useful_life,
                'Start Year': item.start_year
            })
        
        if items_data:
            items_df = pd.DataFrame(items_data)
            st.dataframe(items_df, use_container_width=True, hide_index=True)
        else:
            st.info("No capital assets configured")
        
        # Summary
        col1, col2, col3 = st.columns(3)
        total_capex = st.session_state.capex_manager.total_capex()
        
        with col1:
            st.metric("Total CAPEX", f"${total_capex/1e6:.2f}M")
        with col2:
            st.metric("# Assets", len(assets))
        with col3:
            deprec_schedule = st.session_state.capex_manager.depreciation_schedule(2026, 5)
            st.metric("2026 Depreciation", f"${deprec_schedule.get(2026, 0)/1e3:.0f}K")
    
    # TAB 2: Add Asset
    with tab2:
        st.markdown("## Add New Capital Asset")
        
        col1, col2 = st.columns(2)
        
        with col1:
            name = st.text_input("Asset Name", value="New Asset")
            category = st.text_input("Asset Category", value="Equipment")
            amount = st.number_input("Acquisition Cost ($)", min_value=10000, value=100000, step=10000)
            useful_life = st.number_input("Useful Life (years)", min_value=1, value=10)
        
        with col2:
            salvage_value = st.number_input("Salvage Value ($)", min_value=0, value=0, step=5000)
            start_year = st.number_input("Start Year", min_value=2026, value=2026)
            notes = st.text_area("Notes", value="")
        
        if st.button("✅ Add Asset"):
            try:
                asset_id = st.session_state.capex_manager.add_item(
                    name=name,
                    amount=amount,
                    start_year=int(start_year),
                    useful_life=int(useful_life),
                    salvage_value=salvage_value,
                    category=category,
                    notes=notes
                )
                st.success(f"✅ Asset added! ID: {asset_id}")
                st.session_state.last_update = datetime.now()
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    # TAB 3: Edit Asset
    with tab3:
        st.markdown("## Edit Capital Asset")
        
        assets = st.session_state.capex_manager.list_items()
        if not assets:
            st.warning("No assets to edit")
        else:
            asset_display = {f"{item.item_id} - {item.name}": item.item_id for item in assets}
            selected_display = st.selectbox("Select Asset", list(asset_display.keys()))
            selected_id = asset_display[selected_display]
            asset = st.session_state.capex_manager.get_item(selected_id)
            
            col1, col2 = st.columns(2)
            
            with col1:
                new_name = st.text_input("Name", value=asset.name)
                new_category = st.text_input("Category", value=asset.category)
                new_amount = st.number_input("Amount ($)", min_value=0, value=int(asset.amount))
            
            with col2:
                new_life = st.number_input("Useful Life (years)", min_value=1, value=asset.useful_life)
                new_salvage = st.number_input("Salvage Value", min_value=0, value=int(asset.salvage_value))
                new_notes = st.text_area("Notes", value=asset.notes)
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("✅ Save Changes"):
                    try:
                        st.session_state.capex_manager.edit_item(
                            selected_id,
                            name=new_name,
                            category=new_category,
                            amount=new_amount,
                            useful_life=new_life,
                            salvage_value=new_salvage,
                            notes=new_notes
                        )
                        st.success("✅ Asset updated!")
                        st.session_state.last_update = datetime.now()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
            
            with col2:
                if st.button("🗑️ Remove Asset"):
                    try:
                        st.session_state.capex_manager.remove_item(selected_id)
                        st.success("✅ Asset removed!")
                        st.session_state.last_update = datetime.now()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")

# =====================================================
# PAGE 4: FINANCIAL MODEL
# =====================================================

elif page == "💰 Financial Model":
    st.markdown("# 💰 Financial Model & Valuation")
    
    tab1, tab2 = st.tabs(["🧮 Run Model", "📊 Results"])
    
    with tab1:
        st.markdown("## Model Configuration")
        
        col1, col2 = st.columns(2)
        
        with col1:
            wacc = st.slider("WACC (%)", 0, 20, 8) / 100
            terminal_growth = st.slider("Terminal Growth Rate (%)", 0, 5, 2) / 100
            revenue_cagr = st.slider("Revenue CAGR (%)", 0, 30, 15) / 100
            cogs_percent = st.slider("COGS % of Revenue", 0, 100, 65) / 100
        
        with col2:
            tax_rate = st.slider("Tax Rate (%)", 0, 50, 21) / 100
            debt_amount = st.number_input("Debt ($M)", min_value=0, value=50) * 1e6
            interest_rate = st.slider("Interest Rate (%)", 0, 15, 5) / 100
            shares_outstanding = st.number_input("Shares Outstanding (M)", min_value=1, value=100) * 1e6
        
        if st.button("🧮 Run Financial Model"):
            try:
                cfg = CompanyConfig(
                    revenue_cagr=revenue_cagr,
                    cogs_percent=cogs_percent,
                    tax_rate=tax_rate,
                    wacc=wacc,
                    terminal_growth_rate=terminal_growth,
                    debt_amount=debt_amount,
                    interest_rate=interest_rate,
                    shares_outstanding=shares_outstanding,
                    labor_manager=st.session_state.labor_manager,
                    capex_manager=st.session_state.capex_manager
                )
                
                st.session_state.financial_model = run_financial_model(cfg)
                st.session_state.last_update = datetime.now()
                st.success("✅ Model executed successfully!")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    with tab2:
        if st.session_state.financial_model:
            model = st.session_state.financial_model
            
            st.markdown("## Financial Results")
            
            # Metrics
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Enterprise Value", f"${model['enterprise_value']/1e6:.1f}M")
            with col2:
                st.metric("5-Year FCF", f"${sum(model['fcf'])/1e6:.1f}M")
            with col3:
                ev_revenue = model['enterprise_value'] / model['revenue'][0]
                st.metric("EV/Revenue", f"{ev_revenue:.1f}x")
            with col4:
                st.metric("Terminal Value", f"${model['revenue'][4]*5/1e6:.1f}M")
            
            # Forecast table
            forecast_df = pd.DataFrame({
                'Year': range(2026, 2031),
                'Revenue': [f"${x/1e6:.1f}M" for x in model['revenue']],
                'EBIT': [f"${x/1e6:.1f}M" for x in model['ebit']],
                'FCF': [f"${x/1e6:.1f}M" for x in model['fcf']],
                'Cash': [f"${x/1e6:.1f}M" for x in model['cash_balance']]
            })
            
            st.dataframe(forecast_df, use_container_width=True, hide_index=True)
            
            # Charts
            chart_df = pd.DataFrame({
                'Year': range(2026, 2031),
                'Revenue': model['revenue'],
                'EBIT': model['ebit'],
                'FCF': model['fcf']
            })
            
            fig = px.line(chart_df, x='Year', y=['Revenue', 'EBIT', 'FCF'], markers=True,
                         title="5-Year Forecast")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Run the model to see results")

# =====================================================
# PAGE 5: REPORTS
# =====================================================

elif page == "📈 Reports":
    st.markdown("# 📈 Financial Reports & Exports")
    
    if st.session_state.financial_model:
        model = st.session_state.financial_model
        
        # Summary Report
        st.markdown("## Executive Summary")
        
        summary_text = f"""
        **2026 Financials:**
        - Revenue: ${model['revenue'][0]/1e6:.1f}M
        - EBIT: ${model['ebit'][0]/1e6:.1f}M
        - Net Profit: ${model['net_profit'][0]/1e6:.1f}M
        - FCF: ${model['fcf'][0]/1e6:.1f}M
        
        **Valuation:**
        - Enterprise Value: ${model['enterprise_value']/1e6:.1f}M
        - 5-Year FCF: ${sum(model['fcf'])/1e6:.1f}M
        
        **Workforce:**
        - Headcount: {st.session_state.labor_manager.get_total_headcount(2026)} employees
        - Labor Cost: ${(st.session_state.labor_manager.get_labor_cost_by_type(2026, st.session_state.salary_growth_rate)['Direct'] + st.session_state.labor_manager.get_labor_cost_by_type(2026, st.session_state.salary_growth_rate)['Indirect'])/1e6:.2f}M
        """
        
        st.markdown(summary_text)
        
        # ========== EXCEL MODEL DOWNLOAD SECTION ==========
        st.markdown("## 📊 Excel Model Download")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            selected_scenario = st.selectbox(
                "Select Scenario:",
                ["Base Case", "Conservative", "Optimistic"],
                key="scenario_selector"
            )
            st.session_state.selected_scenario = selected_scenario
        
        download_container = st.container()
        
        # Excel bytes cache management
        excel_map: Dict[str, bytes] = st.session_state.setdefault("excel_bytes_map", {})
        excel_bytes = excel_map.get(selected_scenario)
        
        model.scenario = selected_scenario
        
        with download_container:
            col_prepare, col_download, col_clear = st.columns([2, 2, 2])
            
            with col_prepare:
                if not excel_bytes:
                    if st.button("📄 Prepare Excel Model", key=f"prepare_excel_{selected_scenario.lower()}"):
                        with st.spinner("Preparing Excel workbook..."):
                            excel_bytes = _generate_excel_bytes(
                                model, 
                                st.session_state.labor_manager,
                                st.session_state.capex_manager,
                                selected_scenario
                            )
                        if excel_bytes:
                            excel_map[selected_scenario] = excel_bytes
                            st.session_state.excel_bytes_map = excel_map
                            st.success("✅ Excel model ready for download!")
                        else:
                            st.error("❌ Failed to generate Excel file")
            
            with col_download:
                if excel_bytes:
                    st.download_button(
                        "⬇️ Download Excel Model",
                        data=excel_bytes,
                        file_name=f"Automobile_Manufacturing_Financial_Model_{selected_scenario.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_excel_{selected_scenario.lower()}"
                    )
            
            with col_clear:
                if excel_bytes:
                    if st.button("🗑️ Clear Prepared Excel", key=f"clear_excel_{selected_scenario.lower()}"):
                        excel_map.pop(selected_scenario, None)
                        st.session_state.excel_bytes_map = excel_map
                        excel_bytes = None
                        st.rerun()
        
        if not excel_bytes:
            st.info("💡 Click 'Prepare Excel Model' to generate the workbook for download. It will include Executive Summary, Financial Model, Labor Schedule, and CAPEX Schedule.")
        
        st.divider()
        
        # ========== CSV EXPORTS SECTION ==========
        st.markdown("## 📋 CSV Exports")
        st.markdown("*Quick export of individual reports in CSV format*")
        
        col1, col2 = st.columns(2)
        
        # Financial summary
        with col1:
            export_df = pd.DataFrame({
                'Year': range(2026, 2031),
                'Revenue': model['revenue'],
                'COGS': model['cogs'],
                'OPEX': model['opex'],
                'EBIT': model['ebit'],
                'Net Profit': model['net_profit'],
                'FCF': model['fcf']
            })
            
            csv = export_df.to_csv(index=False)
            st.download_button(
                "📥 Download Financial Forecast (CSV)",
                csv,
                "financial_forecast.csv",
                "text/csv",
                key="download_financial_csv"
            )
        
        # Labor summary
        with col2:
            cost_schedule = LaborCostSchedule(st.session_state.labor_manager)
            labor_df = cost_schedule.generate_5year_schedule(salary_growth=st.session_state.salary_growth_rate)
            
            csv = labor_df.to_csv(index=False)
            st.download_button(
                "📥 Download Labor Schedule (CSV)",
                csv,
                "labor_schedule.csv",
                "text/csv",
                key="download_labor_csv"
            )
    else:
        st.info("Run the financial model first to generate reports")

# =====================================================
# FOOTER
# =====================================================

st.divider()

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("### 📱 Quick Links")
    st.markdown("""
    - [GitHub](https://github.com/Kossit73/Automobile_Manufacturing)
    - [Documentation](https://github.com/Kossit73/Automobile_Manufacturing/blob/main/README.md)
    - [Labor Guide](https://github.com/Kossit73/Automobile_Manufacturing/blob/main/LABOR_MANAGEMENT_GUIDE.md)
    """)

with col2:
    st.markdown("### 📚 Features")
    st.markdown("""
    - ✓ Labor CRUD management
    - ✓ CAPEX scheduling
    - ✓ Financial modeling
    - ✓ DCF valuation
    - ✓ Advanced reports
    """)

with col3:
    st.markdown("### 🔧 Info")
    if st.session_state.last_update:
        st.write(f"Last update: {st.session_state.last_update.strftime('%H:%M:%S')}")
    st.write("Platform v1.0")
    st.write("Python 3.7+")

st.markdown("""
---
**Automobile Manufacturing Financial Platform** © 2025 | [GitHub](https://github.com/Kossit73/Automobile_Manufacturing)
""")
